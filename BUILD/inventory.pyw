import openpyxl as pyxl
from datetime import datetime
import csv

class Inventory(object):
    def __init__(self):
        self.db_filepath = str()
        # NOTES:
        # scanned_data = dict()
        # sorted_valid_items = list()
        # lookup_table = dict()
        # categories = dict()
        self.import_and_parse_success = False
          
    def split_into_chars(self, word:str):
        return [ char for char in word ]

    def parse_scanner_data(self, scanned_data:str):
        self.scanned_data = {}
        for line in scanned_data.splitlines():
            try:
                item = line.split()
                if len(item) == 3:
                    if item[1] in self.scanned_data:
                        temp = self.scanned_data[item[1]][2]
                        item[2] = str(int(temp)+int(item[2]))
                        self.scanned_data[item[1]] = item
                    else:
                        self.scanned_data[item[1]] = item
            except:
                continue
    
    def get_categories(self, valid_items:list):
        categories = set()
        self.categories = dict()
        for item in valid_items:
            categories.add(item[5])
        for category in categories:
            self.categories[category] = 0

    def get_valid_items(self, filename:str) -> list:
        #col 62 is UPC with suffix and prefix
        sku, reg_price, name, stock, upc_code, category, full_upc_code = 0, 7, 1, 11, 12, 13, 62
        valid_items = []
        with open(filename, 'r', encoding='utf-8') as f:
            for line in f.readlines():
                line = line.split('\t')
                try:
                    qty = int(line[11])
                    if line[12].isdigit():
                        if len( self.split_into_chars(line[12]) ) >= 5 :
                            valid_items.append( [ line[sku], line[reg_price], line[name], line[stock], line[upc_code], line[category], line[full_upc_code] ] )
                except:
                    continue
        return valid_items

    def sort_by_categories(self, valid_items:list):
        valid_items.sort(key=lambda x: x[5])
        self.sorted_valid_items = valid_items

    def get_lookup_table(self, valid_items:list) -> dict:
        self.lookup_table = {}
        #pick the biggest upc if available because scanner cannot remove suffix or check digit 
        for item in valid_items:
            if len(item[6]) > len(item[4]):
                self.lookup_table[ item[6] ] = item
            else:
                self.lookup_table[ item[4] ] = item

    def import_database_file(self, db_filepath):
        self.db_filepath = db_filepath
        valid_items = self.get_valid_items(db_filepath)
        self.sort_by_categories(valid_items)
        self.get_lookup_table(valid_items)
        self.get_categories(valid_items)
        self.import_and_parse_success = True
    
    def generate_inventory_report(self, scanned_data:list):
        first_row = ['UPC Code', 'Name', 'Price', 'QTY in POS', 'QTY onhand', 'Difference', 'Category']
        data_for_report = []
        self.parse_scanner_data(scanned_data)

        for item in self.sorted_valid_items:
            sku, reg_price, name, qty_in_pos, upc_code, category, full_upc_code = item[0] ,item[1], item[2], item[3], item[4], item[5], item[6]
            if upc_code in self.scanned_data or full_upc_code in self.scanned_data:
                try:
                    qty_onhand = self.scanned_data[upc_code][2]
                except:
                    qty_onhand = self.scanned_data[full_upc_code][2]

                difference = str( int(qty_onhand) - int(qty_in_pos) )
                if category not in data_for_report:
                    data_for_report.append(category)
                    data_for_report.append( [upc_code, name, reg_price, qty_in_pos, qty_onhand, difference, category] )
                else:
                    data_for_report.append( [upc_code, name, reg_price, qty_in_pos, qty_onhand, difference, category] )
        if data_for_report:
            workbook = pyxl.Workbook()
            sheet = workbook.active
            r = 1
            for line in data_for_report:
                if type(line) is str:
                    r = 2
                    sheet = workbook.create_sheet(line)
                    col = 1
                    for item in first_row:
                        sheet.cell(row=1, column=col).value = item
                        col+=1
                else:
                    c = 1
                    for item in line:
                        if c in [1, 4 ,5, 6]:
                            item = int(item)
                        if c in [3]:
                            item = float(item)
                        sheet.cell(row=r, column=c).value = item
                        c+=1
                    r+=1
            try:
                sht_rem = workbook.get_sheet_by_name('Sheet')
                workbook.remove_sheet(sht_rem)
            except:
                pass
            workbook.save(filename='Inventory_report-' + str(datetime.now().strftime("%Y-%m-%d-%H-%M-%S")) + '.xlsx' )
 
    def generate_unscanned_report(self, scanned_data:list):
        first_row = ['UPC Code', 'Name', 'Price', 'QTY in POS','Category']
        data_for_report = []
        self.parse_scanner_data(scanned_data)

        for item in self.sorted_valid_items:
            sku, reg_price, name, qty_in_pos, upc_code, category, full_upc_code = item[0] ,item[1], item[2], item[3], item[4], item[5], item[6]
            if upc_code not in self.scanned_data and full_upc_code not in self.scanned_data:
                if category not in data_for_report:
                    data_for_report.append(category)
                    data_for_report.append([upc_code, name, reg_price, qty_in_pos, category])
                else:
                    data_for_report.append( [upc_code, name, reg_price, qty_in_pos, category] )
        if data_for_report:
            workbook = pyxl.Workbook()
            sheet = workbook.active
            r = 1
            for line in data_for_report:
                if type(line) is str:
                    r = 2
                    sheet = workbook.create_sheet(line)
                    col = 1
                    for item in first_row:
                        sheet.cell(row=1, column=col).value = item
                        col+=1
                else:
                    c = 1
                    for item in line:
                        if c in [1,4]:
                            item = int(item)
                        if c in [3]:
                            item = float(item)
                        sheet.cell(row=r, column=c).value = item
                        c+=1
                    r+=1
            try:
                sht_rem = workbook.get_sheet_by_name('Sheet')
                workbook.remove_sheet(sht_rem)
            except:
                pass
            workbook.save(filename='Unscanned_report-' + str(datetime.now().strftime("%Y-%m-%d-%H-%M-%S")) + '.xlsx' )

    def generate_reports_of_items_not_in_database(self, scanned_data:list):
        data_for_report = [['', 'UPC Code', 'QTY Scanned']]
        self.parse_scanner_data(scanned_data)

        for item in self.scanned_data:
            if item not in self.sorted_valid_items:
                data_for_report.append(self.scanned_data[item])

        if len(data_for_report) > 1:
            workbook = pyxl.Workbook()
            sheet = workbook.active
            sheet = workbook.create_sheet("Items not in Database")
            r = 1
            for row in data_for_report:
                col = 1
                for item in row:
                    sheet.cell(row=r, column=col).value = item
                    col+=1
                r+=1
            try:
                sht_rem = workbook.get_sheet_by_name('Sheet')
                workbook.remove_sheet(sht_rem)
            except:
                pass
            workbook.save(filename='Products_not_in_database_report-' + str(datetime.now().strftime("%Y-%m-%d-%H-%M-%S")) + '.xlsx' )
    
    def generate_db_file(self, scanned_data):
        sku, reg_price, name, stock, upc_code, category = 0, 7, 1, 11, 12 ,13
        new_data = []
        self.parse_scanner_data(scanned_data)
        with open(self.db_filepath, 'r', encoding='utf-8') as f:
            for line in f.readlines():
                line = line.split('\t')
                if line[upc_code] in self.scanned_data and self.categories[line[category]] == 1:
                    line[stock] = self.scanned_data[line[upc_code]][2]
                    new_data.append(line)
                else:
                    new_data.append(line)
        new_db_filename = 'New_database-' + datetime.now().strftime("%Y-%m-%d-%H-%M-%S")+ '.csv'
        with open(new_db_filename, 'w', encoding='utf-8', newline='') as f:
            csv_writer = csv.writer(f, delimiter='\t' )
            for line in new_data:
                csv_writer.writerow(line[:-1])
                
    def generate_item_import_file(self, scanned_data):
        sku, reg_price, name, stock, upc_code, category, full_upc_code = 0, 7, 1, 11, 12 ,13, 62
        self.parse_scanner_data(scanned_data)
        new_data = []
        with open(self.db_filepath, 'r', encoding='utf-8') as f:
            for line in f.readlines():
                line = line.split('\t')
                if line[upc_code] in self.scanned_data or line[full_upc_code] in self.scanned_data:
                    try:
                        line[stock] = self.scanned_data[line[full_upc_code]][2]
                    except:
                        line[stock] = self.scanned_data[line[upc_code]][2]
                    new_data.append(line)
                else:
                    continue
        new_db_filename = 'Item_Import_file-' + datetime.now().strftime("%Y-%m-%d-%H-%M-%S")+ '.csv'
        with open(new_db_filename, 'w', encoding='utf-8', newline='') as f:
            csv_writer = csv.writer(f, delimiter=',' )
            for line in new_data:
                row  = [line[sku], '0', line[stock] ]
                csv_writer.writerow(row)

    def workbook_formatting(self):
        """Format the excel workbooks to look nice"""
        pass

if __name__ == '__main__':
    inv = Inventory()
    inv.import_database_file('./csv files/br-union.csv')
    inv.categories = {
        'LIQUOR' : 1, 
        'WINE COOLERS' : 0, 
        'SODA' : 0,
        'BEER' : 1, 
        'MISC. NON-TAXABLE' : 0, 
        'CIGAR' : 1, 
        'CHAMPAGNE' : 1, 
        'UNKNOWN' : 0, 
        'WINE' : 1, 
        'LOTTERY' : 0 }
    inv.scanned_data = 'BRUNION\t8218409046\t999\nBRUNION\t8700000737\t999\nBRUNION\t8832000402\t0\nBRUNION\t8811002130\t20\nBRUNION\t4900005010\t12\n'
    # inv.generate_inventory_report(inv.scanned_data)
    inv.generate_item_import_file(inv.scanned_data)
    # inv.generate_db_file()

    print('Done')