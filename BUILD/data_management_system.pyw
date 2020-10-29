from tkinter import *
from tkinter import filedialog
from tkinter.messagebox import showinfo
import openpyxl as pyxl
import sys
import platform
from pathlib import Path
from pathlib import WindowsPath

class Inventory(object):
    def __init__(self):
        db_filepath = ''
        scanned_data = {}
        sorted_valid_items = []
        lookup_table = {}
        categories = [] 
        
    def split_into_chars(self, word:str):
        return [ char for char in word ]

    def parse_scanner_data(self, scanned_data:list):
        self.scanned_data = {}
        for line in scanned_data.splitlines():
            try:
                item = line.split()
                if len(item) == 3:
                    self.scanned_data[item[1]] = item
            except:
                continue
    
    def get_categories(self, valid_items:list):
        categories = set()
        for item in valid_items:
            categories.add(item[5])
        self.categories = list(categories)

    def get_valid_items(self, filename:str) -> list:
        sku, reg_price, name, stock, upc_code, category = 0, 7, 1, 11, 12 ,13
        valid_items = []
        with open(filename, 'r', encoding='utf-8') as f:
            for line in f.readlines():
                line = line.split('\t')
                try:
                    qty = int(line[11])
                    if line[12].isdigit():
                        if len( self.split_into_chars(line[12]) ) >= 5 :
                            valid_items.append( [ line[sku], line[reg_price], line[name], line[stock], line[upc_code], line[category] ])
                except:
                    continue
        return valid_items

    def sort_by_categories(self, valid_items:list):
        valid_items.sort(key=lambda x: x[5])
        self.sorted_valid_items = valid_items

    def get_lookup_table(self, valid_items:list) -> dict:
        self.lookup_table = {}
        for item in valid_items:
            self.lookup_table[ item[4] ] = item

    def import_database_file(self):
        self.db_filepath = filedialog.askopenfilename(initialdir = ".",title = "Select file",filetypes = (("csv files","*.csv"),("all files","*.*")))
        if self.db_filepath is not None:
            valid_items = self.get_valid_items(self.db_filepath)
            self.sort_by_categories(valid_items)
            self.get_lookup_table(valid_items)


def generate_inventory_report(inventory:Inventory, scanned_data:list):
    #change to xlsx with different tabs for each category
    # upc_code : sku, reg_price, name, stock, upc_code, category
    first_row = ['UPC Code', 'Name', 'Price', 'QTY in POS', 'QTY onhand', 'Difference', 'Category']
    data_for_report = []
    inventory.parse_scanner_data(scanned_data)

    for item in inventory.sorted_valid_items:
        sku, reg_price, name, qty_in_pos, upc_code, category = item[0] ,item[1], item[2], item[3], item[4], item[5]
        if upc_code in inventory.scanned_data:
            qty_onhand = inventory.scanned_data[upc_code][2]
            difference = str( int(qty_onhand) - int(qty_in_pos) )
            if category not in data_for_report:
                data_for_report.append(category)
                data_for_report.append([upc_code, name, reg_price, qty_in_pos, qty_onhand, difference, category])
            else:
                data_for_report.append( [upc_code, name, reg_price, qty_in_pos, qty_onhand, difference, category] )
    if data_for_report:
        workbook = pyxl.Workbook()
        r = 1
        for line in data_for_report:
            if type(line) is str:
                r = 2
                sheet = workbook.create_sheet(line)
                col = 1
                for item in first_row:
                    sheet.cell(row=1, column=col).value = item
                    col+=1
            else:
                c = 1
                for item in line:
                    if c in [1, 4 ,5, 6]:
                        item = int(item)
                    if c in [3]:
                        item = float(item)
                    sheet.cell(row=r, column=c).value = item
                    c+=1
                r+=1
        try:
            sht_rem = workbook.get_sheet_by_name('Sheet')
            workbook.remove_sheet(sht_rem)
        except:
            pass
        workbook.save(filename='inventory_report.xlsx')

def generate_unscanned_report(inventory:Inventory, scanned_data:list):
    # categories = ['BEER', 'CHAMPAGNE', 'CIGAR', 'LIQUOR', 'MISC. NON-TAXABLE', 'MISC. TAXABLE', 'SODA', 'UNKNOWN', 'WINE', 'WINE COOLERS'] #col 13
    # scanned_data = inventory.parse_scanner_data(scanned_data)
    pass
    
def generate_new_database(filename:str):
    pass

if __name__ == "__main__":
    inventory = Inventory()
    root = Tk()
    root.title('Cellar 97 Inventory Management')
    # root.iconbitmap(default='titan.ico')
    root.minsize(795, 490)
    root.maxsize(795, 490)
    menubar = Menu(root)
    filemenu = Menu(menubar, tearoff=0)
    # filemenu.add_command(label="Help", command=help)
    filemenu.add_command(label="Open Log File Folder")
    filemenu.add_separator()
    filemenu.add_command(label="Close", command=root.destroy)
    menubar.add_cascade(label="File", menu=filemenu)
    try:
        if platform.system() == 'Windows':
            banner = WindowsPath(sys._MEIPASS, 'banner.png')
            icon = WindowsPath(sys._MEIPASS, 'cellar97.ico')
            background_image = PhotoImage(file=banner)
            root.iconbitmap(default=icon)
    except:
        background_image = PhotoImage(file='banner.png')
        root.iconbitmap(default='cellar97.ico')
    
    background_label = Label(root, image=background_image, borderwidth=0, highlightthickness=0)
    background_label.place( relx=0.00, rely=0.00)
    version_label = Label(text='Version 1.0', bg='#000000' ,fg='#ffffff').place(relx=0.915, rely=0.16)
    root.config(menu=menubar, bg='#c2c6cc')

    scanner_label = Label(text='Insert scanned data below:', bg='#c2c6cc' ,fg='#000000').place(relx=0.15, rely=0.209)

    left_frame = Frame(root, bg='#dae2f0')
    left_frame.place(relx=0.005 ,rely=0.25)
    left_frame_scrollbar = Scrollbar(left_frame)
    left_frame_scrollbar.pack(side=RIGHT, fill=Y)
    left_frame_text = Text(left_frame, yscrollcommand=left_frame_scrollbar.set,  width=50, height=22 , bg='#dae2f0')
    left_frame_text.pack()
    left_frame_scrollbar.config(command=left_frame_text.yview)

    import_button = Button(root, text='Import Current DB', command=lambda:inventory.import_database_file())
    import_button.place(relx=0.68, rely=0.28)
    
    gen_inventory_button = Button(root, text='Generate Inventory Report', command=lambda:generate_inventory_report(inventory, left_frame_text.get('1.0', END)))
    gen_inventory_button.place(relx=0.68, rely=0.36)
    
    gen_inventory_button = Button(root, text='Generate Unscanned Items Report')
    gen_inventory_button.place(relx=0.68, rely=0.44)

    gen_inventory_button = Button(root, text='Generate New DB')
    gen_inventory_button.place(relx=0.68, rely=0.52)

    
    root.mainloop()

    # hashmap = csv_to_hashmap_filtered("./csv files/br-union.csv")
    # print(hashmap['3125900145'])

    
    print("done")