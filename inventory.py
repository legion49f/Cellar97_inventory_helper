import openpyxl as pyxl

class Inventory(object):
    def __init__(self):
        self.db_filepath = ''
        scanned_data = {}
        sorted_valid_items = []
        lookup_table = {}
        categories = [] 
        
    def split_into_chars(self, word:str):
        return [ char for char in word ]

    def parse_scanner_data(self, scanned_data:list):
        self.scanned_data = {}
        for line in scanned_data.splitlines():
            try:
                item = line.split()
                if len(item) == 3:
                    self.scanned_data[item[1]] = item
            except:
                continue
    
    def get_categories(self, valid_items:list):
        categories = set()
        for item in valid_items:
            categories.add(item[5])
        self.categories = list(categories)

    def get_valid_items(self, filename:str) -> list:
        sku, reg_price, name, stock, upc_code, category = 0, 7, 1, 11, 12 ,13
        valid_items = []
        with open(filename, 'r', encoding='utf-8') as f:
            for line in f.readlines():
                line = line.split('\t')
                try:
                    qty = int(line[11])
                    if line[12].isdigit():
                        if len( self.split_into_chars(line[12]) ) >= 5 :
                            valid_items.append( [ line[sku], line[reg_price], line[name], line[stock], line[upc_code], line[category] ])
                except:
                    continue
        return valid_items

    def sort_by_categories(self, valid_items:list):
        valid_items.sort(key=lambda x: x[5])
        self.sorted_valid_items = valid_items

    def get_lookup_table(self, valid_items:list) -> dict:
        self.lookup_table = {}
        for item in valid_items:
            self.lookup_table[ item[4] ] = item

    def import_database_file(self, db_filepath):
        valid_items = self.get_valid_items(db_filepath)
        self.sort_by_categories(valid_items)
        self.get_lookup_table(valid_items)
    
    def generate_inventory_report(self, scanned_data:list):
        first_row = ['UPC Code', 'Name', 'Price', 'QTY in POS', 'QTY onhand', 'Difference', 'Category']
        data_for_report = []
        self.parse_scanner_data(scanned_data)

        for item in self.sorted_valid_items:
            sku, reg_price, name, qty_in_pos, upc_code, category = item[0] ,item[1], item[2], item[3], item[4], item[5]
            if upc_code in self.scanned_data:
                qty_onhand = self.scanned_data[upc_code][2]
                difference = str( int(qty_onhand) - int(qty_in_pos) )
                if category not in data_for_report:
                    data_for_report.append(category)
                    data_for_report.append([upc_code, name, reg_price, qty_in_pos, qty_onhand, difference, category])
                else:
                    data_for_report.append( [upc_code, name, reg_price, qty_in_pos, qty_onhand, difference, category] )
        if data_for_report:
            workbook = pyxl.Workbook()
            sheet = workbook.active
            r = 1
            for line in data_for_report:
                if type(line) is str:
                    r = 2
                    sheet = workbook.create_sheet(line)
                    col = 1
                    for item in first_row:
                        sheet.cell(row=1, column=col).value = item
                        col+=1
                else:
                    c = 1
                    for item in line:
                        if c in [1, 4 ,5, 6]:
                            item = int(item)
                        if c in [3]:
                            item = float(item)
                        sheet.cell(row=r, column=c).value = item
                        c+=1
                    r+=1
            try:
                sht_rem = workbook.get_sheet_by_name('Sheet')
                workbook.remove_sheet(sht_rem)
            except:
                pass
            workbook.save(filename='Inventory_report.xlsx')

    def generate_unscanned_report(self, scanned_data:list):
        first_row = ['UPC Code', 'Name', 'Price', 'QTY in POS','Category']
        data_for_report = []
        self.parse_scanner_data(scanned_data)

        for item in self.sorted_valid_items:
            sku, reg_price, name, qty_in_pos, upc_code, category = item[0] ,item[1], item[2], item[3], item[4], item[5]
            if upc_code not in self.scanned_data:
                if category not in data_for_report:
                    data_for_report.append(category)
                    data_for_report.append([upc_code, name, reg_price, qty_in_pos, category])
                else:
                    data_for_report.append( [upc_code, name, reg_price, qty_in_pos, category] )
        if data_for_report:
            workbook = pyxl.Workbook()
            sheet = workbook.active
            r = 1
            for line in data_for_report:
                if type(line) is str:
                    r = 2
                    sheet = workbook.create_sheet(line)
                    col = 1
                    for item in first_row:
                        sheet.cell(row=1, column=col).value = item
                        col+=1
                else:
                    c = 1
                    for item in line:
                        if c in [1,4]:
                            item = int(item)
                        if c in [3]:
                            item = float(item)
                        sheet.cell(row=r, column=c).value = item
                        c+=1
                    r+=1
            try:
                sht_rem = workbook.get_sheet_by_name('Sheet')
                workbook.remove_sheet(sht_rem)
            except:
                pass
            workbook.save(filename='Unscanned_report.xlsx')
    
    def generate_db_file(self):
        pass

    def workbook_formatting(self):
        pass